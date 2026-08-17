"""
Baixador unificado de documentos fiscais - NF-e, CT-e e NFS-e
=============================================================
Varre os tres ambientes nacionais por NSU usando o certificado A1 e grava
os XMLs em uma unica pasta de saida (entrada do classificador de vencimentos).

  NF-e  -> NFeDistribuicaoDFe   (SOAP)  distNSU
  CT-e  -> CTeDistribuicaoDFe   (SOAP)  distNSU
  NFS-e -> ADN /contribuintes/DFe/{NSU} (REST) paginacao por NSU

Configuracao em config.ini (ao lado deste arquivo). Estado (ultNSU por
servico, bloqueios) em PASTA_CONTROLE - nunca no OneDrive.

Requisitos:
    pip install requests cryptography openpyxl

Uso:
    python baixador.py                       # varre os tres servicos
    python baixador.py --servico nfe cte     # so NF-e e CT-e
    python baixador.py --status              # mostra o estado e sai
    python baixador.py --servico nfse --nsu 0    # reprocessa do inicio
"""

import os
import re
import sys
import ssl
import csv
import json
import gzip
import time
import base64
import logging
import argparse
import tempfile
import configparser
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path

import requests
from requests.adapters import HTTPAdapter
from cryptography.hazmat.primitives.serialization import (
    pkcs12, Encoding, PrivateFormat, NoEncryption,
)

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


log = logging.getLogger("baixador")

SERVICOS = ("nfe", "cte", "nfse")

NS_NFE = "http://www.portalfiscal.inf.br/nfe"
NS_CTE = "http://www.portalfiscal.inf.br/cte"

URL_NFE = {
    "1": "https://www1.nfe.fazenda.gov.br/NFeDistribuicaoDFe/NFeDistribuicaoDFe.asmx",
    "2": "https://hom1.nfe.fazenda.gov.br/NFeDistribuicaoDFe/NFeDistribuicaoDFe.asmx",
}
URL_CTE = {
    "1": "https://www1.cte.fazenda.gov.br/CTeDistribuicaoDFe/CTeDistribuicaoDFe.asmx",
    "2": "https://hom1.cte.fazenda.gov.br/CTeDistribuicaoDFe/CTeDistribuicaoDFe.asmx",
}
# Ambiente de Dados Nacional da NFS-e (distribuicao por NSU)
URL_NFSE = {
    "1": "https://adn.nfse.gov.br",
    "2": "https://adn.producaorestrita.nfse.gov.br",
}
# Sefin Nacional da NFS-e (consulta avulsa por chave de acesso)
URL_SEFIN_NFSE = {
    "1": "https://sefin.nfse.gov.br/SefinNacional",
    "2": "https://sefin.producaorestrita.nfse.gov.br/SefinNacional",
}

ACTION_NFE = "http://www.portalfiscal.inf.br/nfe/wsdl/NFeDistribuicaoDFe/nfeDistDFeInteresse"
ACTION_CTE = "http://www.portalfiscal.inf.br/cte/wsdl/CTeDistribuicaoDFe/cteDistDFeInteresse"

# cStat do DistribuicaoDFe que interessam ao controle da varredura
CSTAT_NENHUM_DOCUMENTO = "137"   # nenhum documento localizado
CSTAT_DOCUMENTO_LOCALIZADO = "138"
CSTAT_CONSUMO_INDEVIDO = "656"   # bloqueio de 1h imposto pela SEFAZ

# StatusProcessamento devolvido pelo ADN da NFS-e
NFSE_SEM_DOCUMENTO = "NENHUM_DOCUMENTO_LOCALIZADO"
NFSE_REJEICAO = "REJEICAO"

# Raizes que representam o documento fiscal em si (o resto e evento/resumo)
RAIZES_DOCUMENTO = {
    "nfeProc", "NFe",
    "cteProc", "CTeOSProc", "cteSimpProc", "CTe",
    "nfseProc", "NFSe",
}

# Id="NFe3524...", Id="CTe...", Id="NFS..." ou a propria chave solta
RE_CHAVE_ID = re.compile(r'Id="(?:NFe|CTe|NFS|DPS)?(\d{44,50})"')
RE_CHAVE_SOLTA = re.compile(r"(?<!\d)(\d{44,50})(?!\d)")

# Tags de data, em ordem de preferencia: a primeira encontrada define a data
# do documento (emissao quando existe, senao autorizacao/registro).
TAGS_DATA = (
    "dhEmi", "dEmi", "dhProc", "dhRecbto",
    "dhEvento", "dhRegEvento", "DataEmissao", "dhProcessamento",
)
RE_DATA_ISO = re.compile(r"(\d{4})-(\d{2})-(\d{2})")

# Modelo do documento: posicoes 21-22 da chave de 44 digitos
MODELO_SERVICO = {"55": "nfe", "65": "nfe", "57": "cte", "67": "cte"}
MAX_CHAVES = 10


# ===================== CONFIGURACAO =====================
def base_execucao():
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent


@dataclass
class Config:
    cnpj: str
    cuf: str
    tp_amb: str = "1"
    cert_pfx: Path = None
    cert_senha: str = ""
    pasta_saida: Path = None
    pasta_controle: Path = None
    pasta_planilhas: Path = None
    servicos: dict = field(default_factory=lambda: dict.fromkeys(SERVICOS, True))
    max_lotes: int = 20
    pausa: int = 2
    bloqueio_min: int = 65
    timeout: int = 90
    periodo_de: date = None
    periodo_ate: date = None

    @property
    def arquivo_estado(self):
        return self.pasta_controle / "estado_nsu.json"

    @property
    def arquivo_lock(self):
        return self.pasta_controle / "baixador.lock"

    @property
    def arquivo_relatorio(self):
        return self.pasta_controle / "relatorio_faltantes.csv"

    @property
    def arquivo_log(self):
        return self.pasta_controle / "baixador.log"

    @classmethod
    def de_arquivo(cls, caminho):
        cfg = configparser.ConfigParser()
        lidos = cfg.read(caminho, encoding="utf-8")
        if not lidos:
            raise FileNotFoundError(f"Nao consegui ler o config: {caminho}")

        def obrig(secao, chave):
            valor = cfg.get(secao, chave, fallback="").strip()
            if not valor:
                raise ValueError(f"config.ini: preencha [{secao}] {chave}")
            return valor

        cnpj = re.sub(r"\D", "", obrig("EMPRESA", "cnpj"))
        if len(cnpj) != 14:
            raise ValueError("config.ini: [EMPRESA] cnpj deve ter 14 digitos")

        tp_amb = cfg.get("EMPRESA", "ambiente", fallback="1").strip()
        if tp_amb not in ("1", "2"):
            raise ValueError("config.ini: [EMPRESA] ambiente deve ser 1 (producao) ou 2 (homologacao)")

        base = base_execucao()
        pasta_saida = Path(obrig("PASTAS", "saida")).expanduser()
        pasta_controle = Path(obrig("PASTAS", "controle")).expanduser()
        pasta_planilhas = Path(cfg.get("PASTAS", "planilhas", fallback=str(base)).strip() or base).expanduser()

        ligado = {"sim", "true", "1", "s", "yes"}
        servicos = {
            s: cfg.get("SERVICOS", s, fallback="sim").strip().lower() in ligado
            for s in SERVICOS
        }

        def inteiro(secao, chave, padrao):
            """getint proprio: a opcao existir vazia no .ini nao pode estourar."""
            bruto = cfg.get(secao, chave, fallback="").strip()
            if not bruto:
                return padrao
            try:
                return int(bruto)
            except ValueError:
                raise ValueError(
                    f"config.ini: [{secao}] {chave} deve ser um numero ({bruto!r})"
                ) from None

        periodo_de = ler_data(cfg.get("PERIODO", "de", fallback=""))
        periodo_ate = ler_data(cfg.get("PERIODO", "ate", fallback=""))
        ultimos = inteiro("PERIODO", "ultimos_dias", 0)
        if ultimos and not periodo_de:
            periodo_de = date.today() - timedelta(days=ultimos)

        return cls(
            cnpj=cnpj,
            cuf=obrig("EMPRESA", "cuf"),
            tp_amb=tp_amb,
            cert_pfx=Path(obrig("CERTIFICADO", "pfx")).expanduser(),
            cert_senha=cfg.get("CERTIFICADO", "senha", fallback=""),
            pasta_saida=pasta_saida,
            pasta_controle=pasta_controle,
            pasta_planilhas=pasta_planilhas,
            servicos=servicos,
            max_lotes=inteiro("LIMITES", "max_lotes_por_execucao", 20),
            pausa=inteiro("LIMITES", "pausa_segundos", 2),
            bloqueio_min=inteiro("LIMITES", "bloqueio_minutos", 65),
            timeout=inteiro("LIMITES", "timeout_segundos", 90),
            periodo_de=periodo_de,
            periodo_ate=periodo_ate,
        )

    def validar_periodo(self):
        if self.periodo_de and self.periodo_ate and self.periodo_de > self.periodo_ate:
            raise ValueError("periodo invalido: a data inicial e maior que a final")

    def descricao_periodo(self):
        if not self.periodo_de and not self.periodo_ate:
            return "sem filtro (tudo que o ambiente nacional entregar)"
        de = self.periodo_de.strftime("%d/%m/%Y") if self.periodo_de else "o inicio"
        ate = self.periodo_ate.strftime("%d/%m/%Y") if self.periodo_ate else "hoje"
        return f"de {de} ate {ate}"


def ler_data_iso(texto):
    """Primeira data aaaa-mm-dd encontrada no texto (ex.: campo JSON do ADN)."""
    m = RE_DATA_ISO.search(texto or "")
    if not m:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None


def ler_data(texto):
    """Aceita dd/mm/aaaa, dd-mm-aaaa e aaaa-mm-dd. Vazio -> None."""
    texto = (texto or "").strip()
    if not texto:
        return None
    for formato in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    raise ValueError(f"data invalida: {texto} (use dd/mm/aaaa)")


def configurar_log(cfg, verboso=False):
    log.setLevel(logging.DEBUG if verboso else logging.INFO)
    log.handlers.clear()

    console = logging.StreamHandler(sys.stdout)
    console.setFormatter(logging.Formatter("%(message)s"))
    log.addHandler(console)

    try:
        cfg.pasta_controle.mkdir(parents=True, exist_ok=True)
        arquivo = logging.FileHandler(cfg.arquivo_log, encoding="utf-8")
        arquivo.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
        log.addHandler(arquivo)
    except Exception as e:  # log em arquivo e conveniencia, nao requisito
        print(f"AVISO: sem log em arquivo ({e})")


# ===================== LOCK =====================
def adquirir_lock(cfg):
    cfg.pasta_controle.mkdir(parents=True, exist_ok=True)
    if cfg.arquivo_lock.exists():
        try:
            dados = json.loads(cfg.arquivo_lock.read_text(encoding="utf-8"))
            inicio = datetime.fromisoformat(dados["inicio"])
            if datetime.now() - inicio < timedelta(hours=2):
                log.info("Execucao em andamento (PID %s). Saindo.", dados.get("pid"))
                return False
            log.info("Lock antigo (>2h), assumindo travado.")
        except Exception:
            pass
    cfg.arquivo_lock.write_text(
        json.dumps({"pid": os.getpid(), "inicio": datetime.now().isoformat()}),
        encoding="utf-8",
    )
    return True


def liberar_lock(cfg):
    try:
        cfg.arquivo_lock.unlink(missing_ok=True)
    except Exception:
        pass


# ===================== ESTADO =====================
def carregar_estado(cfg):
    if cfg.arquivo_estado.exists():
        try:
            return json.loads(cfg.arquivo_estado.read_text(encoding="utf-8"))
        except Exception:
            log.warning("Estado corrompido, recriando do zero.")
    return {}


def salvar_estado(cfg, estado):
    cfg.arquivo_estado.parent.mkdir(parents=True, exist_ok=True)
    temporario = cfg.arquivo_estado.with_suffix(".tmp")
    temporario.write_text(
        json.dumps(estado, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    temporario.replace(cfg.arquivo_estado)


def estado_servico(estado, servico):
    return estado.setdefault(
        servico, {"ultNSU": "0", "maxNSU": "0", "bloqueado_ate": None}
    )


def bloqueado(info):
    if not info.get("bloqueado_ate"):
        return None
    try:
        ate = datetime.fromisoformat(info["bloqueado_ate"])
        return ate if datetime.now() < ate else None
    except Exception:
        return None


def bloquear(cfg, info, motivo):
    ate = datetime.now() + timedelta(minutes=cfg.bloqueio_min)
    info["bloqueado_ate"] = ate.isoformat()
    info["motivo_bloqueio"] = motivo
    log.warning("    !! BLOQUEADO ate %s (%s)", ate.strftime("%d/%m %H:%M"), motivo)


def como_inteiro(valor, padrao=0):
    try:
        return int(str(valor).strip())
    except (TypeError, ValueError):
        return padrao


# ===================== CERTIFICADO =====================
class TLSAdapter(HTTPAdapter):
    """Alguns endpoints da SEFAZ exigem renegociacao legada no OpenSSL 3."""

    def init_poolmanager(self, *args, **kwargs):
        ctx = ssl.create_default_context()
        ctx.check_hostname = True
        try:
            ctx.options |= 0x4  # OP_LEGACY_SERVER_CONNECT
        except Exception:
            pass
        ctx.set_ciphers("DEFAULT@SECLEVEL=1")
        kwargs["ssl_context"] = ctx
        return super().init_poolmanager(*args, **kwargs)


def preparar_certificado(cfg):
    if not cfg.cert_pfx.exists():
        raise FileNotFoundError(f"Certificado nao encontrado: {cfg.cert_pfx}")
    try:
        chave, cert, cadeia = pkcs12.load_key_and_certificates(
            cfg.cert_pfx.read_bytes(), cfg.cert_senha.encode("utf-8")
        )
    except ValueError as e:
        raise ValueError(f"Nao consegui abrir o .pfx (senha errada?): {e}") from e
    if cert is None or chave is None:
        raise ValueError("Nao consegui extrair certificado/chave do .pfx")

    validade = getattr(cert, "not_valid_after_utc", None) or cert.not_valid_after
    validade = validade.replace(tzinfo=None)
    agora = datetime.utcnow()
    if validade < agora:
        raise ValueError(f"Certificado VENCIDO em {validade:%d/%m/%Y}")
    if (validade - agora).days < 15:
        log.warning("AVISO: certificado vence em %s", validade.strftime("%d/%m/%Y"))

    fd_c, pem_cert = tempfile.mkstemp(suffix=".pem")
    fd_k, pem_key = tempfile.mkstemp(suffix=".pem")
    with os.fdopen(fd_c, "wb") as f:
        f.write(cert.public_bytes(Encoding.PEM))
        for extra in (cadeia or []):
            f.write(extra.public_bytes(Encoding.PEM))
    with os.fdopen(fd_k, "wb") as f:
        f.write(chave.private_bytes(Encoding.PEM, PrivateFormat.PKCS8, NoEncryption()))
    return pem_cert, pem_key


def nova_sessao(cert_par):
    s = requests.Session()
    s.cert = cert_par
    s.mount("https://", TLSAdapter())
    return s


# ===================== GRAVACAO =====================
def raiz_do_xml(conteudo):
    """Nome da tag raiz, sem namespace. String vazia se o XML for ilegivel."""
    try:
        return ET.fromstring(conteudo).tag.split("}")[-1]
    except ET.ParseError:
        return ""


def extrair_chave(conteudo):
    """Chave de acesso (44 digitos NF-e/CT-e, 50 na NFS-e), ou None."""
    m = RE_CHAVE_ID.search(conteudo) or RE_CHAVE_SOLTA.search(conteudo)
    return m.group(1) if m else None


def nome_arquivo(conteudo, prefixo, raiz=None, nsu=None, documento=True):
    """
    Documento fiscal: <chave>.xml - permite deduplicar por chave.
    Evento/resumo: <chave>-<raiz>-<nsu>.xml - varios eventos por chave,
    entao o nome precisa de discriminante ou um sobrescreveria o outro.
    """
    chave = extrair_chave(conteudo)
    if documento and chave:
        return f"{chave}.xml"

    raiz = raiz or raiz_do_xml(conteudo) or prefixo
    sufixo = str(nsu).strip() if nsu else datetime.now().strftime("%Y%m%d%H%M%S%f")
    base = chave or prefixo
    return f"{base}-{raiz}-{sufixo}.xml"


def data_do_documento(conteudo):
    """Data do documento (emissao de preferencia), ou None se nao achar."""
    try:
        raiz = ET.fromstring(conteudo)
    except ET.ParseError:
        return None

    achados = {}
    for el in raiz.iter():
        tag = el.tag.split("}")[-1]
        if tag in TAGS_DATA and el.text and tag not in achados:
            achados[tag] = el.text

    for tag in TAGS_DATA:
        data = ler_data_iso(achados.get(tag, ""))
        if data:
            return data
    return None


def dentro_do_periodo(cfg, data):
    """Sem periodo configurado, tudo entra. Sem data legivel, tambem entra:
    melhor gravar um XML a mais do que perder documento por falha de leitura."""
    if data is None:
        return True
    if cfg.periodo_de and data < cfg.periodo_de:
        return False
    if cfg.periodo_ate and data > cfg.periodo_ate:
        return False
    return True


def gravar_xml(cfg, conteudo, prefixo, contadores, nsu=None, data=None,
               ignorar_periodo=False):
    """Grava documentos processaveis na raiz; eventos/resumos em subpasta."""
    if not ignorar_periodo:
        if data is None:
            data = data_do_documento(conteudo)
        if not dentro_do_periodo(cfg, data):
            contadores["fora_periodo"] += 1
            return None

    raiz = raiz_do_xml(conteudo)
    documento = raiz in RAIZES_DOCUMENTO

    if documento:
        pasta = cfg.pasta_saida
        contadores["documentos"] += 1
    else:
        pasta = cfg.pasta_saida / "_EVENTOS_E_RESUMOS"
        contadores["eventos"] += 1

    pasta.mkdir(parents=True, exist_ok=True)
    destino = pasta / nome_arquivo(conteudo, prefixo, raiz, nsu, documento)
    if destino.exists():
        contadores["documentos" if documento else "eventos"] -= 1
        contadores["duplicados"] += 1
        return None
    destino.write_text(conteudo, encoding="utf-8")
    return destino


def descompactar_doczip(texto_b64):
    """docZip da SEFAZ / ArquivoXml do ADN: base64 de gzip (as vezes XML puro)."""
    bruto = base64.b64decode(texto_b64)
    try:
        bruto = gzip.decompress(bruto)
    except (OSError, EOFError):
        pass
    return bruto.decode("utf-8", errors="replace").lstrip("\ufeff")


# ===================== MODULO SOAP (NF-e / CT-e) =====================
@dataclass
class RespostaDist:
    """retDistDFeInt ja interpretado: docs = [(nsu, xml, data)]."""
    cstat: str = ""
    xmotivo: str = ""
    ult_nsu: str = ""
    max_nsu: str = ""
    docs: list = field(default_factory=list)

    @property
    def datas(self):
        return [d for _, _, d in self.docs if d]


def montar_envelope(cfg, servico, ult_nsu=None, chave=None):
    """distNSU para varredura; consChNFe/consChCTe para consulta avulsa."""
    if servico == "nfe":
        ns = NS_NFE
        wsdl = "http://www.portalfiscal.inf.br/nfe/wsdl/NFeDistribuicaoDFe"
        metodo = "nfeDistDFeInteresse"
        tag_dados = "nfeDadosMsg"
        versao = "1.01"
        consulta_chave = f"<consChNFe><chNFe>{chave}</chNFe></consChNFe>"
    else:
        ns = NS_CTE
        wsdl = "http://www.portalfiscal.inf.br/cte/wsdl/CTeDistribuicaoDFe"
        metodo = "cteDistDFeInteresse"
        tag_dados = "cteDadosMsg"
        versao = "1.00"
        consulta_chave = f"<consChCTe><chCTe>{chave}</chCTe></consChCTe>"

    if chave:
        corpo = consulta_chave
    else:
        corpo = f"<distNSU><ultNSU>{str(ult_nsu).zfill(15)}</ultNSU></distNSU>"

    return (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<soap12:Envelope xmlns:soap12="http://www.w3.org/2003/05/soap-envelope">'
        "<soap12:Body>"
        f'<{metodo} xmlns="{wsdl}">'
        f"<{tag_dados}>"
        f'<distDFeInt xmlns="{ns}" versao="{versao}">'
        f"<tpAmb>{cfg.tp_amb}</tpAmb>"
        f"<cUFAutor>{cfg.cuf}</cUFAutor>"
        f"<CNPJ>{cfg.cnpj}</CNPJ>"
        f"{corpo}"
        "</distDFeInt>"
        f"</{tag_dados}>"
        f"</{metodo}>"
        "</soap12:Body>"
        "</soap12:Envelope>"
    )


def consultar_soap(cfg, servico, sessao, ult_nsu=None, chave=None):
    """Uma requisicao ao DistribuicaoDFe. None quando a resposta nao veio."""
    ns = NS_NFE if servico == "nfe" else NS_CTE
    url = (URL_NFE if servico == "nfe" else URL_CTE)[cfg.tp_amb]
    action = ACTION_NFE if servico == "nfe" else ACTION_CTE

    try:
        resp = sessao.post(
            url,
            data=montar_envelope(cfg, servico, ult_nsu, chave).encode("utf-8"),
            headers={
                "Content-Type": f'application/soap+xml; charset=utf-8; action="{action}"'
            },
            timeout=cfg.timeout,
        )
        resp.raise_for_status()
    except Exception as e:
        log.error("    falha de rede: %s", e)
        return None

    try:
        raiz = ET.fromstring(resp.content)
    except ET.ParseError as e:
        log.error("    resposta ilegivel: %s", e)
        return None

    def txt(tag):
        el = raiz.find(f".//{{{ns}}}{tag}")
        return el.text if el is not None and el.text else ""

    resposta = RespostaDist(
        cstat=txt("cStat"),
        xmotivo=txt("xMotivo"),
        ult_nsu=txt("ultNSU"),
        max_nsu=txt("maxNSU"),
    )
    for doc in raiz.findall(f".//{{{ns}}}docZip"):
        try:
            xml = descompactar_doczip(doc.text)
        except Exception as e:
            log.error("    erro ao descompactar NSU %s: %s", doc.get("NSU"), e)
            continue
        resposta.docs.append((doc.get("NSU") or "", xml, data_do_documento(xml)))
    return resposta


def avancar_ate_periodo(cfg, servico, sessao, info, margem=200, max_sondagens=20):
    """
    Busca binaria pelo NSU onde o periodo comeca, para nao baixar os lotes
    anteriores. O NSU cresce junto com a chegada dos documentos ao ambiente
    nacional, entao da para sondar. A margem cobre documento autorizado com
    atraso, que entra com NSU maior que o da sua data de emissao.
    """
    if not cfg.periodo_de:
        return

    atual = como_inteiro(info["ultNSU"])
    resposta = consultar_soap(cfg, servico, sessao, atual)
    if resposta is None:
        return
    if resposta.cstat == CSTAT_CONSUMO_INDEVIDO:
        bloquear(cfg, info, f"{resposta.cstat} {resposta.xmotivo}")
        return
    if resposta.datas and min(resposta.datas) >= cfg.periodo_de:
        log.info("  o NSU atual ja esta dentro do periodo; nada a pular.")
        return

    baixo, alto = atual, como_inteiro(resposta.max_nsu)
    if alto <= baixo:
        return

    sondagens = 0
    while alto - baixo > margem and sondagens < max_sondagens:
        meio = (baixo + alto) // 2
        time.sleep(cfg.pausa)
        resposta = consultar_soap(cfg, servico, sessao, meio)
        sondagens += 1
        if resposta is None:
            break
        if resposta.cstat == CSTAT_CONSUMO_INDEVIDO:
            bloquear(cfg, info, f"{resposta.cstat} {resposta.xmotivo}")
            return

        datas = resposta.datas
        log.info(
            "  sondagem NSU %s -> %s",
            meio,
            max(datas).strftime("%d/%m/%Y") if datas else "sem documento",
        )
        if datas and max(datas) < cfg.periodo_de:
            baixo = meio
        else:
            alto = meio

    novo = max(baixo - margem, atual)
    if novo > atual:
        log.info(
            "  pulando do NSU %s para %s (%s sondagem(ns)); os documentos"
            " anteriores nao serao baixados.", atual, novo, sondagens,
        )
        info["ultNSU"] = str(novo)
    else:
        log.info("  nao houve NSU a pular.")


def rodar_soap(cfg, servico, sessao, info, contadores):
    prefixo = servico.upper()

    for lote in range(1, cfg.max_lotes + 1):
        ult = info["ultNSU"]
        log.info("  lote %s/%s a partir do NSU %s", lote, cfg.max_lotes, ult)

        resposta = consultar_soap(cfg, servico, sessao, ult)
        if resposta is None:
            return
        log.info(
            "    cStat %s - %s | %s doc(s)",
            resposta.cstat, resposta.xmotivo, len(resposta.docs),
        )

        if resposta.cstat == CSTAT_CONSUMO_INDEVIDO:
            bloquear(cfg, info, f"{resposta.cstat} {resposta.xmotivo}")
            return

        for nsu, xml, data in resposta.docs:
            try:
                gravar_xml(cfg, xml, prefixo, contadores, nsu=nsu, data=data)
            except Exception as e:
                log.error("    erro ao gravar NSU %s: %s", nsu, e)

        if resposta.ult_nsu:
            info["ultNSU"] = resposta.ult_nsu
        if resposta.max_nsu:
            info["maxNSU"] = resposta.max_nsu

        if not resposta.docs:
            if resposta.cstat == CSTAT_NENHUM_DOCUMENTO:
                log.info("    nada novo, varredura concluida.")
            elif resposta.cstat != CSTAT_DOCUMENTO_LOCALIZADO:
                log.warning("    parando: cStat %s - %s", resposta.cstat, resposta.xmotivo)
            return
        if como_inteiro(resposta.ult_nsu) >= como_inteiro(resposta.max_nsu, -1) > 0:
            log.info("    alcancou o maxNSU, varredura concluida.")
            return
        if como_inteiro(resposta.ult_nsu) <= como_inteiro(ult):
            log.warning("    NSU nao avancou (%s); interrompendo para nao repetir.", ult)
            return
        if cfg.periodo_ate and resposta.datas and min(resposta.datas) > cfg.periodo_ate:
            log.info(
                "    o lote ja passou de %s; parando por aqui.",
                cfg.periodo_ate.strftime("%d/%m/%Y"),
            )
            return
        time.sleep(cfg.pausa)

    log.info(
        "    limite de %s lote(s) atingido; continua na proxima execucao.", cfg.max_lotes
    )


# ===================== MODULO REST (NFS-e / ADN) =====================
def rodar_nfse(cfg, sessao, info, contadores):
    """
    ADN NFS-e: GET /contribuintes/DFe/{ultimo NSU processado}?lote=true
    Devolve LoteDFe com ArquivoXml (base64 de gzip) e o NSU de cada item.
    Nao ha maxNSU: para quando o lote vier vazio ou HTTP 404.
    """
    base_url = URL_NFSE[cfg.tp_amb]

    for lote in range(1, cfg.max_lotes + 1):
        ult = como_inteiro(info["ultNSU"])
        log.info("  lote %s/%s a partir do NSU %s", lote, cfg.max_lotes, ult)
        try:
            resp = sessao.get(
                f"{base_url}/contribuintes/DFe/{ult}",
                params={"lote": "true", "cnpjConsulta": cfg.cnpj},
                headers={"Accept": "application/json"},
                timeout=cfg.timeout,
            )
        except Exception as e:
            log.error("    falha de rede: %s", e)
            return

        if resp.status_code == 404:
            log.info("    nada novo, varredura concluida.")
            return
        if resp.status_code in (401, 403):
            log.error("    acesso negado (%s). Confira se o CNPJ do", resp.status_code)
            log.error("    certificado e ator (tomador) das notas e se o ADN esta habilitado.")
            return
        if resp.status_code == 429:
            bloquear(cfg, info, "429 too many requests")
            return
        if resp.status_code >= 400:
            log.error("    HTTP %s: %s", resp.status_code, resp.text[:300])
            return

        try:
            dados = resp.json()
        except ValueError:
            log.error("    resposta nao-JSON: %s", resp.text[:300])
            return

        documentos = dados.get("LoteDFe") or dados.get("loteDFe") or []
        status = dados.get("StatusProcessamento") or dados.get("statusProcessamento") or ""
        log.info("    %s | %s documento(s)", status or "sem status", len(documentos))

        for erro in dados.get("Erros") or []:
            log.error("    erro ADN: %s", erro)

        maior_nsu = ult
        datas = []
        for item in documentos:
            conteudo_b64 = item.get("ArquivoXml") or item.get("arquivoXml")
            nsu_item = como_inteiro(item.get("NSU") or item.get("nsu"), 0)
            maior_nsu = max(maior_nsu, nsu_item)
            if not conteudo_b64:
                continue
            try:
                xml = descompactar_doczip(conteudo_b64)
                data = data_do_documento(xml) or ler_data_iso(item.get("DataHoraGeracao"))
                if data:
                    datas.append(data)
                gravar_xml(cfg, xml, "NFSE", contadores, nsu=nsu_item or None, data=data)
            except Exception as e:
                log.error("    erro ao processar NSU %s: %s", nsu_item, e)

        if maior_nsu > ult:
            info["ultNSU"] = str(maior_nsu)
            info["maxNSU"] = str(maior_nsu)

        if status == NFSE_REJEICAO:
            log.error("    ADN rejeitou a consulta; interrompendo.")
            return
        if not documentos or status == NFSE_SEM_DOCUMENTO:
            log.info("    nada novo, varredura concluida.")
            return
        if maior_nsu <= ult:
            log.warning("    NSU nao avancou (%s); interrompendo para nao repetir.", ult)
            return
        if cfg.periodo_ate and datas and min(datas) > cfg.periodo_ate:
            log.info(
                "    o lote ja passou de %s; parando por aqui.",
                cfg.periodo_ate.strftime("%d/%m/%Y"),
            )
            return
        time.sleep(cfg.pausa)

    log.info(
        "    limite de %s lote(s) atingido; continua na proxima execucao.", cfg.max_lotes
    )


# ===================== CONSULTA AVULSA POR CHAVE =====================
def servico_da_chave(chave):
    """
    Descobre o servico pela propria chave: 50 digitos e NFS-e; nos 44 digitos
    o modelo esta nas posicoes 21-22 (55/65 = NF-e, 57/67 = CT-e).
    """
    if len(chave) == 50:
        return "nfse"
    if len(chave) == 44:
        return MODELO_SERVICO.get(chave[20:22])
    return None


def normalizar_chaves(entradas):
    """
    Aceita chave crua, com espacos em grupos de 4 (como o portal mostra) ou
    varias separadas por virgula/ponto-e-virgula/quebra de linha.
    Devolve (chaves_validas, entradas_recusadas).
    """
    validas, erros = [], []

    def guardar(chave):
        if chave not in validas:
            validas.append(chave)

    texto = " ".join(str(e) for e in entradas)
    for candidato in re.split(r"[,;\n]+", texto):
        candidato = candidato.strip()
        if not candidato:
            continue
        pedacos = candidato.split()

        # Se algum pedaco ja e uma chave inteira, sao varias chaves soltas -
        # juntar os digitos aqui poderia formar uma chave que ninguem pediu.
        if len(pedacos) > 1 and any(servico_da_chave(re.sub(r"\D", "", p)) for p in pedacos):
            for pedaco in pedacos:
                digitos = re.sub(r"\D", "", pedaco)
                if servico_da_chave(digitos):
                    guardar(digitos)
                else:
                    erros.append(pedaco)
            continue

        # Senao, e uma chave so - possivelmente em grupos de 4 digitos
        digitos = re.sub(r"\D", "", candidato)
        if servico_da_chave(digitos):
            guardar(digitos)
        else:
            erros.append(candidato if len(candidato) <= 60 else candidato[:57] + "...")

    return validas, erros


def baixar_nfse_por_chave(cfg, sessao, chave, contadores):
    """Sefin Nacional: GET /nfse/{chave} devolve o XML em base64/gzip."""
    url = f"{URL_SEFIN_NFSE[cfg.tp_amb]}/nfse/{chave}"
    try:
        resp = sessao.get(url, headers={"Accept": "application/json"}, timeout=cfg.timeout)
    except Exception as e:
        log.error("    falha de rede: %s", e)
        return False

    if resp.status_code == 404:
        log.warning("    nao encontrada no Sefin Nacional.")
        return False
    if resp.status_code >= 400:
        log.error("    HTTP %s: %s", resp.status_code, resp.text[:300])
        return False

    try:
        dados = resp.json()
    except ValueError:
        log.error("    resposta nao-JSON: %s", resp.text[:300])
        return False

    bruto = dados.get("nfseXmlGZipB64") or dados.get("NfseXmlGZipB64")
    if bruto:
        xml = descompactar_doczip(bruto)
    else:
        xml = dados.get("xmlNFSe") or dados.get("xml") or ""
        if not xml.lstrip().startswith("<"):
            log.warning("    resposta sem XML da NFS-e.")
            return False

    destino = gravar_xml(cfg, xml, "NFSE", contadores, ignorar_periodo=True)
    log.info("    %s", destino.name if destino else "ja estava na pasta.")
    return True


def baixar_soap_por_chave(cfg, servico, sessao, chave, contadores):
    resposta = consultar_soap(cfg, servico, sessao, chave=chave)
    if resposta is None:
        return False
    log.info("    cStat %s - %s", resposta.cstat, resposta.xmotivo)

    if resposta.cstat == CSTAT_CONSUMO_INDEVIDO:
        log.error("    consumo indevido: aguarde ~1h antes de repetir a consulta.")
        return False
    if not resposta.docs:
        log.warning("    nada devolvido (o CNPJ do certificado e parte na nota?).")
        return False

    for nsu, xml, _ in resposta.docs:
        destino = gravar_xml(cfg, xml, servico.upper(), contadores,
                             nsu=nsu, ignorar_periodo=True)
        log.info("    %s", destino.name if destino else "ja estava na pasta.")
    return True


def baixar_por_chave(cfg, sessao, chaves, contadores):
    """Baixa documentos avulsos, sem mexer no NSU nem no filtro de periodo."""
    rotulos = {"nfe": "NF-e", "cte": "CT-e", "nfse": "NFS-e"}
    achados = 0

    for i, chave in enumerate(chaves, 1):
        servico = servico_da_chave(chave)
        log.info("\n[%s/%s] %s %s", i, len(chaves), rotulos[servico], chave)
        try:
            if servico == "nfse":
                ok = baixar_nfse_por_chave(cfg, sessao, chave, contadores)
            else:
                ok = baixar_soap_por_chave(cfg, servico, sessao, chave, contadores)
        except Exception as e:
            log.exception("    erro inesperado: %s", e)
            ok = False
        achados += bool(ok)
        if i < len(chaves):
            time.sleep(cfg.pausa)

    log.info("\n---\n%s de %s chave(s) atendida(s).", achados, len(chaves))
    return achados


# ===================== CONFERENCIA COM A PLANILHA =====================
def conferir_planilhas(cfg):
    if load_workbook is None:
        log.info("\nConferencia pulada (openpyxl nao instalado).")
        return
    if not cfg.pasta_planilhas.exists():
        return
    arquivos = [c for c in cfg.pasta_planilhas.glob("*.xls*") if not c.name.startswith("~$")]
    if not arquivos:
        return

    baixadas = {p.stem for p in cfg.pasta_saida.glob("*.xml")}
    faltantes = []

    for caminho in arquivos:
        try:
            wb = load_workbook(caminho, read_only=True, data_only=True)
        except Exception as e:
            log.warning("  nao consegui ler %s: %s", caminho.name, e)
            continue
        try:
            ws = wb[wb.sheetnames[0]]
            linhas = ws.iter_rows(values_only=True)
            cab = next(linhas, None)
            if cab is None:
                continue
            idx = {str(n).strip(): i for i, n in enumerate(cab) if n is not None}
            if "Chave" not in idx:
                continue

            def campo(l, n):
                i = idx.get(n)
                return "" if i is None or i >= len(l) or l[i] is None else str(l[i]).strip()

            for linha in linhas:
                if not linha or not any(v is not None and str(v).strip() for v in linha):
                    continue
                chave = re.sub(r"\D", "", campo(linha, "Chave"))
                if len(chave) not in (44, 50) or chave in baixadas:
                    continue
                faltantes.append([
                    caminho.name, chave, campo(linha, "Tipo"), campo(linha, "Num"),
                    campo(linha, "DtAut"), campo(linha, "Valor"), campo(linha, "Emissor Nome"),
                ])
        finally:
            wb.close()

    if not faltantes:
        log.info("\nConferencia: todas as chaves das planilhas estao na pasta de saida.")
        return

    with open(cfg.arquivo_relatorio, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["planilha", "chave", "tipo", "numero", "data_aut", "valor", "emissor"])
        w.writerows(faltantes)
    log.warning("\nConferencia: %s chave(s) da planilha NAO chegaram.", len(faltantes))
    log.info("Detalhes em: %s", cfg.arquivo_relatorio)
    log.info("Causa comum: documento fora da janela de 90 dias do Ambiente Nacional.")


# ===================== MAIN =====================
def montar_argumentos(argv=None):
    p = argparse.ArgumentParser(
        description="Baixa XML de NF-e, CT-e e NFS-e pelos ambientes nacionais."
    )
    p.add_argument("--config", type=Path, default=base_execucao() / "config.ini",
                   help="caminho do config.ini")
    p.add_argument("--servico", nargs="+", choices=SERVICOS, metavar="SERVICO",
                   help=f"limita a varredura ({', '.join(SERVICOS)})")
    p.add_argument("--nsu", type=int,
                   help="forca o NSU inicial (exige um unico --servico)")
    p.add_argument("--chave", nargs="+", metavar="CHAVE",
                   help=f"baixa ate {MAX_CHAVES} documentos por chave de acesso "
                        "(NF-e, CT-e ou NFS-e) e sai; nao mexe no NSU")
    p.add_argument("--de", metavar="DATA",
                   help="so grava documentos a partir desta data (dd/mm/aaaa)")
    p.add_argument("--ate", metavar="DATA",
                   help="so grava documentos ate esta data (dd/mm/aaaa)")
    p.add_argument("--ultimos", type=int, metavar="DIAS",
                   help="atalho para --de = hoje menos DIAS")
    p.add_argument("--pular-anteriores", action="store_true",
                   help="sonda o NSU onde o periodo comeca e pula os lotes "
                        "anteriores (NF-e/CT-e); os antigos nao serao baixados")
    p.add_argument("--max-lotes", type=int, help="sobrepoe o limite do config.ini")
    p.add_argument("--ignorar-bloqueio", action="store_true",
                   help="ignora o bloqueio gravado no estado (cuidado com o 656)")
    p.add_argument("--sem-conferencia", action="store_true",
                   help="nao confere as planilhas no fim")
    p.add_argument("--status", action="store_true",
                   help="mostra o estado atual (NSU/bloqueios) e sai")
    p.add_argument("-v", "--verbose", action="store_true", help="log detalhado")
    return p.parse_args(argv)


def mostrar_status(cfg):
    estado = carregar_estado(cfg)
    log.info("Estado em %s", cfg.arquivo_estado)
    for servico in SERVICOS:
        info = estado.get(servico)
        ligado = "ligado" if cfg.servicos.get(servico) else "desligado"
        if not info:
            log.info("  %-5s [%s] nunca executado", servico, ligado)
            continue
        ate = bloqueado(info)
        trava = f" | BLOQUEADO ate {ate:%d/%m %H:%M}" if ate else ""
        log.info("  %-5s [%s] ultNSU=%s maxNSU=%s%s",
                 servico, ligado, info.get("ultNSU"), info.get("maxNSU"), trava)


def main(argv=None):
    args = montar_argumentos(argv)

    if not args.config.exists():
        print(f"config.ini nao encontrado em {args.config}")
        print("Copie o config.ini.exemplo, ajuste os caminhos e rode de novo.")
        return 1
    try:
        cfg = Config.de_arquivo(args.config)
    except (ValueError, FileNotFoundError, configparser.Error) as e:
        print(f"ERRO no config.ini: {e}")
        return 1

    if args.max_lotes:
        cfg.max_lotes = args.max_lotes

    try:
        if args.de:
            cfg.periodo_de = ler_data(args.de)
        if args.ate:
            cfg.periodo_ate = ler_data(args.ate)
        if args.ultimos:
            cfg.periodo_de = date.today() - timedelta(days=args.ultimos)
        cfg.validar_periodo()
    except ValueError as e:
        print(f"ERRO: {e}")
        return 1

    configurar_log(cfg, args.verbose)

    if args.status:
        mostrar_status(cfg)
        log.info("Periodo configurado: %s", cfg.descricao_periodo())
        return 0

    chaves = erros_chave = None
    if args.chave:
        chaves, erros_chave = normalizar_chaves(args.chave)
        for ruim in erros_chave:
            print(f"chave invalida (esperado 44 ou 50 digitos): {ruim}")
        if not chaves:
            return 1
        if len(chaves) > MAX_CHAVES:
            print(f"informe no maximo {MAX_CHAVES} chaves por vez ({len(chaves)} recebidas)")
            return 1

    escolhidos = args.servico or [s for s in SERVICOS if cfg.servicos.get(s)]
    if args.nsu is not None and len(escolhidos) != 1:
        print("--nsu exige exatamente um --servico")
        return 1
    if not chaves and not escolhidos:
        log.info("Nenhum servico ativo no config.ini.")
        return 0

    cfg.pasta_saida.mkdir(parents=True, exist_ok=True)
    cfg.pasta_controle.mkdir(parents=True, exist_ok=True)

    if not adquirir_lock(cfg):
        return 1

    pem_cert = pem_key = None
    try:
        try:
            pem_cert, pem_key = preparar_certificado(cfg)
        except Exception as e:
            log.error("ERRO no certificado: %s", e)
            return 1

        sessao = nova_sessao((pem_cert, pem_key))
        contadores = {"documentos": 0, "eventos": 0, "duplicados": 0, "fora_periodo": 0}

        if chaves:
            log.info("Consulta avulsa de %s chave(s).", len(chaves))
            baixar_por_chave(cfg, sessao, chaves, contadores)
            log.info("Pasta de saida: %s", cfg.pasta_saida)
            return 0

        estado = carregar_estado(cfg)
        log.info("Periodo: %s", cfg.descricao_periodo())

        rotulos = {"nfe": "NF-e", "cte": "CT-e", "nfse": "NFS-e"}
        executores = {
            "nfe": lambda i, c: rodar_soap(cfg, "nfe", sessao, i, c),
            "cte": lambda i, c: rodar_soap(cfg, "cte", sessao, i, c),
            "nfse": lambda i, c: rodar_nfse(cfg, sessao, i, c),
        }

        for servico in escolhidos:
            info = estado_servico(estado, servico)
            log.info("\n== %s ==", rotulos[servico])

            if args.nsu is not None:
                log.info("  NSU inicial forcado para %s", args.nsu)
                info["ultNSU"] = str(args.nsu)

            ate = bloqueado(info)
            if ate and not args.ignorar_bloqueio:
                restante = int((ate - datetime.now()).total_seconds() // 60) + 1
                log.info("  bloqueado por ~%s min (ate %s). Pulando.", restante, ate.strftime("%H:%M"))
                continue
            info["bloqueado_ate"] = None
            info.pop("motivo_bloqueio", None)

            try:
                if args.pular_anteriores and servico in ("nfe", "cte"):
                    avancar_ate_periodo(cfg, servico, sessao, info)
                    salvar_estado(cfg, estado)
                    if bloqueado(info):
                        continue
                elif args.pular_anteriores:
                    log.info("  --pular-anteriores nao vale para NFS-e (o ADN nao"
                             " informa o maxNSU); a varredura segue normal.")
                elif cfg.periodo_de and como_inteiro(info["ultNSU"]) == 0:
                    log.info("  dica: --pular-anteriores evita baixar os lotes"
                             " anteriores a %s.", cfg.periodo_de.strftime("%d/%m/%Y"))

                executores[servico](info, contadores)
            except Exception as e:
                log.exception("  erro inesperado: %s", e)
            finally:
                salvar_estado(cfg, estado)

        salvar_estado(cfg, estado)
        log.info(
            "\n---\nDocumentos gravados: %s | eventos/resumos: %s | ja existentes: %s"
            " | fora do periodo: %s",
            contadores["documentos"], contadores["eventos"],
            contadores["duplicados"], contadores["fora_periodo"],
        )
        log.info("Pasta de saida: %s", cfg.pasta_saida)

        if not args.sem_conferencia:
            conferir_planilhas(cfg)
        return 0
    finally:
        for caminho in (pem_cert, pem_key):
            if caminho:
                try:
                    os.unlink(caminho)
                except Exception:
                    pass
        liberar_lock(cfg)


if __name__ == "__main__":
    sys.exit(main())
